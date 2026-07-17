"""
Parses TG_kalkyl_och_valuta_updated.xlsx -> structured JSON for the quote builder.

Sheet 'Aktuell prislista' structure:
  - Repeating blocks: group header row (col A = group name, col C = 'List price'),
    then a column-header row (tier labels in E:H or E:H, sometimes split across 2 rows),
    then an 'MOQ' row with numeric MOQ per tier,
    then product rows: PartNumber | Description | ListPrice | (EAU/blank) | tier1..tier4 | NegotiationMargin | LifeCycleComment
  - A simpler 'Accessories' / 'Cables' section: PartNo | Description | Price Euro (3 cols only)

Sheet 'TG-kalkyl' has a worked example for ONE article (CCpilot VI, C000 144-05) with:
  SP (sales price SEK), MK (manufacturing cost SEK), Malad TG (target margin %),
  Moms %, exchange rate SEK->EUR.
  We extract MK for that one article as a seed; MK is otherwise unknown per-article.
"""
import openpyxl
import json
import re

SRC = '/mnt/user-data/uploads/TG_kalkyl_och_valuta_updated.xlsx'

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['Aktuell prislista']

TIER_COLS = ['E', 'F', 'G', 'H']  # up to 4 volume tiers
COL_IDX = {'A': 1, 'B': 2, 'C': 3, 'D': 4, 'E': 5, 'F': 6, 'G': 7, 'H': 8, 'I': 9, 'J': 10}

def cell(r, col_letter):
    return ws.cell(r, COL_IDX[col_letter]).value

def is_group_header(r):
    """A group header row: col A has a product-family name, and 'List price'
    appears in col C either on this row OR the next row (split-header case,
    e.g. 'Ccpilot V1090')."""
    a = cell(r, 'A')
    if not (a and isinstance(a, str)):
        return False
    if a.strip() in ('Part number', 'Part No.', 'Accessories', 'Cables'):
        return False
    c_this = cell(r, 'C')
    c_next = cell(r + 1, 'C')
    return c_this == 'List price' or c_next == 'List price'

def is_accessory_section_header(r):
    a = cell(r, 'A')
    return a in ('Accessories', 'Cables')

groups = []
accessory_sections = []

r = 1
max_row = ws.max_row

while r <= max_row:
    a = cell(r, 'A')

    if is_group_header(r):
        group_name = a.strip()
        # tier labels: on this row (E:H) normally, or on next row in split-header case
        tier_labels = [cell(r, c) for c in TIER_COLS]
        header_row = r
        if all(t is None for t in tier_labels):
            header_row = r + 1
            tier_labels = [cell(header_row, c) for c in TIER_COLS]

        # find MOQ row: search next couple rows for D == 'MOQ'
        moq_row = None
        for look in range(r, r + 4):
            if cell(look, 'D') == 'MOQ':
                moq_row = look
                break
        tier_moqs = [cell(moq_row, c) if moq_row else None for c in TIER_COLS]

        # data starts after moq_row (or after header_row+1 fallback)
        data_start = (moq_row + 1) if moq_row else (header_row + 1)

        products = []
        rr = data_start
        while rr <= max_row:
            pn = cell(rr, 'A')
            desc = cell(rr, 'B')
            if pn is None and desc is None:
                # blank line -> end of this group's product block
                break
            if is_group_header(rr) or is_accessory_section_header(rr):
                break
            if isinstance(pn, str) and pn.strip() in ('Part number', 'Part No.'):
                rr += 1
                continue
            list_price = cell(rr, 'C')
            if pn and list_price is not None:
                tiers = {}
                for label, moq, col in zip(tier_labels, tier_moqs, TIER_COLS):
                    val = cell(rr, col)
                    if label and val is not None:
                        tiers[str(label)] = {'moq': moq, 'price_eur': round(float(val), 2)}
                neg_margin = cell(rr, 'I')
                lifecycle = cell(rr, 'J')
                products.append({
                    'part_number': str(pn).strip(),
                    'description': str(desc).strip() if desc else '',
                    'list_price_eur': round(float(list_price), 2),
                    'tiers': tiers,
                    'negotiation_margin': neg_margin if isinstance(neg_margin, (int, float)) else None,
                    'lifecycle_comment': lifecycle if lifecycle else None,
                    'mk_sek': None,  # manufacturing cost - filled from TG-kalkyl seed below if matched
                })
            rr += 1

        if products:
            groups.append({'group': group_name, 'products': products})
        r = rr
        continue

    if is_accessory_section_header(r):
        section_name = a.strip()
        # header row with Part No / Description / Price Euro is 2 rows down typically
        hr = r + 1
        while hr <= max_row and cell(hr, 'A') not in ('Part No.', 'Part number'):
            hr += 1
        rr = hr + 1
        items = []
        while rr <= max_row:
            pn = cell(rr, 'A')
            desc = cell(rr, 'B')
            price = cell(rr, 'C')
            if pn is None and desc is None:
                break
            if is_group_header(rr) or is_accessory_section_header(rr):
                break
            if pn and price is not None:
                items.append({
                    'part_number': str(pn).strip(),
                    'description': str(desc).strip() if desc else '',
                    'price_eur': round(float(price), 2),
                })
            rr += 1
        if items:
            accessory_sections.append({'section': section_name, 'items': items})
        r = rr
        continue

    r += 1

# ---- TG-kalkyl seed: extract the one worked MK example ----
tg = wb['TG-kalkyl']
seed_part_number = tg['A2'].value
seed_mk_sek = tg['B6'].value  # Tillverkningskostnad (MK), SEK
seed_target_tg = tg['B7'].value  # Malad TG (target gross margin %)
seed_vat = tg['B8'].value
seed_fx_sek_eur = tg['C9'].value  # SEK -> EUR rate (mid-market example), note B9 had #VALUE! error

if seed_part_number:
    seed_part_number = str(seed_part_number).strip()
    for g in groups:
        for p in g['products']:
            if p['part_number'] == seed_part_number:
                p['mk_sek'] = round(float(seed_mk_sek), 2)

output = {
    'meta': {
        'currency': 'EUR',
        'source_file': 'TG_kalkyl_och_valuta_updated.xlsx',
        'tg_assumptions': {
            'target_tg': seed_target_tg,
            'vat': seed_vat,
            'fx_sek_eur': seed_fx_sek_eur,
            'note': 'MK (manufacturing cost, SEK) currently only known for one seed article. TG/margin display should only appear when mk_sek is present on a product.'
        }
    },
    'groups': groups,
    'accessories': accessory_sections,
}

with open('/home/claude/pricelist.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"Groups: {len(groups)}")
for g in groups:
    print(f"  {g['group']}: {len(g['products'])} products, tiers={list(g['products'][0]['tiers'].keys()) if g['products'] else []}")
print(f"Accessory sections: {len(accessory_sections)}")
for s in accessory_sections:
    print(f"  {s['section']}: {len(s['items'])} items")
print(f"Seed MK: {seed_part_number} -> {seed_mk_sek} SEK")
